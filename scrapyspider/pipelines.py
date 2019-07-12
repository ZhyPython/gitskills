# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import Workbook
from scrapyspider import items
import json
import pymysql


class ScrapyspiderPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['电影名称', '分数', '评论人数'])

        self.wb_ajax = Workbook()
        self.ws_ajax = self.wb_ajax.active
        self.ws_ajax.append(['排名', '电影名称', '分数', '评论人数'])

    def process_item(self, item, spider):
        if isinstance(item, items.DoubanMovieItem):
            line = [item['movie_name'], item['score'], item['score_num']]
            self.ws.append(line)
            self.wb.save('E:\\scrapyspider\\scrapyspider\\job.xlsx')
            return item

        if isinstance(item, items.DoubanMovieAjax):
            line = [item['ranking'], item['movie_name'], item['score'], item['score_num']]
            self.ws_ajax.append(line)
            self.wb_ajax.save('E:\\scrapyspider\\scrapyspider\\job_ajax.xlsx')
            return item


class JsonPipeline(object):
    """
    process_item处理函数是循环调用的，在spider中，每yield一次item，就调用一次处理函数将数据保存在本地
    """
    def process_item(self, item, spider):
        '''
        保存为json
        :param item:
        :param spider:
        :return:
        '''
        with open('E:\\scrapyspider\\scrapyspider\\ajax.json', 'a', encoding='utf-8') as f:
            line = json.dumps(dict(item), ensure_ascii=False) + "\n"
            f.write(line)
        return item


class MysqlPipeline(object):
    def process_item(self, item, spider):
        movie_name = item['movie_name']
        score = item['score']
        score_num = item['score_num']

        # 连接数据库
        connection = pymysql.connect(
            host='localhost',   # 连接本地数据库
            user='root',
            passwd='123',
            db='scrapydb',
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor)

        try:
            with connection.cursor() as cursor:
                sql = 'INSERT INTO movies (movie_name, score, score_num) VALUES (%s, %s, %s)'
                cursor.execute(sql, (movie_name, score, score_num))
            connection.commit()
        finally:
            connection.close()

        return item


class ProxyPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['地址', 'IP'])

    def process_item(self, item, spider):
        if spider.name == 'kdlspider':
            line = [str(item['addr']), item['port']]
            self.ws.append(line)
            self.wb.save('E:\\scrapyspider\\scrapyspider\\proxy.xlsx')
            return item
